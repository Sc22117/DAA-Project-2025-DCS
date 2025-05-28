import firebase_admin
from firebase_admin import credentials, db
import pandas as pd
import random
from datetime import datetime, timedelta
from collections import defaultdict


# Only initialize the app if it hasn't been initialized already
if not firebase_admin._apps:
    cred = credentials.Certificate("DCS(Firebase Key).json")
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://dcs-simulation---testing-default-rtdb.asia-southeast1.firebasedatabase.app/'
    })


# Reading data from excel

import firebase_admin
from firebase_admin import credentials, db
import pandas as pd
from collections import Counter

def sanitize_key(key: str) -> str:
    return str(key).strip().replace('.', '_').replace('$', '_')\
        .replace('#', '_').replace('[', '_').replace(']', '_')\
        .replace('/', '_').replace(':', '_')

def upload_university_data(institute):
    file_path = "4th Sem Details.xlsx"

    if db.reference(f"{institute}/faculty").get():
        print(f"\n⚠️ Data for '{institute}' already exists in Firebase.")
        generate_excel_reports(institute, output_file="./All_Sections_Report.xlsx")
        exit()

    # Load Excel
    data = {
        'faculty': pd.read_excel(file_path, sheet_name='Faculty'),
        'subjects': pd.read_excel(file_path, sheet_name='Subjects'),
        'sections': pd.read_excel(file_path, sheet_name='Sections'),
        'venues': pd.read_excel(file_path, sheet_name='Venues')
    }

    # Firebase references
    faculty_regular_ref = db.reference(f"{institute}/faculty/regular_subjects_faculty")
    faculty_elective_ref = db.reference(f"{institute}/faculty/electives_faculty")
    subject_regular_ref = db.reference(f"{institute}/subjects/regular_subjects")
    subject_elective_ref = db.reference(f"{institute}/subjects/elective_subjects")
    section_ref = db.reference(f"{institute}/sections")
    venue_ref = db.reference(f"{institute}/venues")

    # --- Build subject name -> subject_code map ---
    subject_code_map = {
        row['Subject'].strip(): sanitize_key(row['Subject Code'])
        for _, row in data['subjects'].iterrows()
    }

    # --- Build reverse map: subject_code -> list of subject names ---
    subject_code_to_names = {}
    for _, row in data['subjects'].iterrows():
        code = sanitize_key(row['Subject Code'])
        name = row['Subject'].strip()
        subject_code_to_names.setdefault(code, []).append(name)

    # --- FACULTY & ELECTIVES ---
    faculty_data = {}
    electives_data = {}

    for _, row in data['faculty'].iterrows():
        faculty_id = sanitize_key(str(row['ID']).strip())
        name = str(row['Name']).strip()

        subjects_raw = [s.strip() for s in str(row['Subject']).split(',')]
        types_raw = [t.strip() for t in str(row['Type']).split(',')]

        for i, subject_raw in enumerate(subjects_raw):
            subj_type = types_raw[i] if i < len(types_raw) else "Theory"
            is_elective = 'Elective' in subject_raw

            subject_original = subject_raw.strip()

            # Clean subject name only for matching
            subject_key_for_match = (
                subject_original.replace(': Elective', '')
                .replace('Elective', '')
                .strip()
            )

            # Try exact match
            subject_code = subject_code_map.get(subject_key_for_match)

            # If not found, try case-insensitive + normalized match
            if not subject_code:
                for sub_name, code in subject_code_map.items():
                    normalized = sub_name.replace(': Elective', '').replace('Elective', '').strip()
                    if subject_key_for_match.lower() == normalized.lower():
                        subject_code = code
                        break

            if not subject_code:
                print(f"⚠️ Subject code not found for: {subject_original}")
                continue

            # --- Use multiple names logic ---
            matched_names = subject_code_to_names.get(subject_code, [subject_original])

            if len(matched_names) > 1:
                # Multiple names for same code → use full subject_original to differentiate
                key_subject = subject_original
            else:
                # Single name → use that
                key_subject = matched_names[0]

            key = sanitize_key(f"{faculty_id}_{key_subject}")

            faculty_entry = {
                'name': name,
                'subject': subject_original,
                'subject_code': subject_code,
                'type': subj_type
            }

            if is_elective:
                electives_data[key] = faculty_entry
            else:
                faculty_data[key] = faculty_entry

    # Upload faculty and electives separately
    faculty_regular_ref.set(faculty_data)
    faculty_elective_ref.set(electives_data)

    # --- SUBJECTS ---
    # Count subject code occurrences
    codes = [sanitize_key(row['Subject Code']) for _, row in data['subjects'].iterrows()]
    code_counts = Counter(codes)

    regular_subjects_data = {}
    elective_subjects_data = {}

    for _, row in data['subjects'].iterrows():
        code = sanitize_key(row['Subject Code'])
        subj_name = row['Subject'].strip()
        entry = {
            'subject': subj_name,
            'type': row['Type'],
            'credits': int(row['Credits'])
        }

        # Make key unique if code occurs multiple times
        key = code
        if code_counts[code] > 1:
            key = f"{code}_{sanitize_key(subj_name)}"

        if 'elective' in subj_name.lower():
            elective_subjects_data[key] = entry
        else:
            regular_subjects_data[key] = entry

    subject_regular_ref.set(regular_subjects_data)
    subject_elective_ref.set(elective_subjects_data)

    # --- SECTIONS ---
    sections_data = {
        sanitize_key(row['Section']): {
            'strength': int(row['Strength']),
            'allocations': {},
            'timetable': {}
        }
        for _, row in data['sections'].iterrows()
    }
    section_ref.set(sections_data)

    # --- VENUES ---
    venues_data = {}
    for _, row in data['venues'].iterrows():
        venue = str(row['Classroom/Lab']).strip()
        if not venue or pd.isna(venue):
            continue
        venues_data[sanitize_key(venue)] = {
            'capacity': int(row['Capacity']),
            'type': 'Lab' if 'Lab' in venue else 'Classroom'
        }
    venue_ref.set(venues_data)

    print(f"\n✅ Data for '{institute}' uploaded successfully.")
    allocate_faculty_to_sections(institute)
    # allocate_classrooms_to_sections(institute)

# Allocating classrooms randomly to each section

# def allocate_classrooms_to_sections(institute):
#     institute_ref = db.reference(f"{institute}")
#     sections_data = institute_ref.child("sections").get()
#     venues_data = institute_ref.child("venues").get()

#     if not sections_data:
#         print("No sections data found under the institute.")
#         return
#     if not venues_data:
#         print("No venues data found under the institute.")
#         return

#     # Filter only classrooms (ignore labs)
#     classrooms = {
#         name: venue['capacity']
#         for name, venue in venues_data.items()
#         if venue.get('type') == 'Classroom'
#     }

#     if not classrooms:
#         print("No classrooms available (all are labs?)")
#         return

#     # Sort classrooms by capacity (smallest to largest)
#     sorted_classrooms = sorted(classrooms.items(), key=lambda x: x[1])
#     used_rooms = set()

#     for section, details in sections_data.items():
#         strength = details.get('strength')
#         if strength is None:
#             print(f"No strength info for section {section}")
#             continue

#         # Get the smallest suitable unused classroom
#         suitable_rooms = [
#             room for room, cap in sorted_classrooms
#             if cap >= strength and room not in used_rooms
#         ]

#         if not suitable_rooms:
#             print(f"⚠️ No suitable classroom for section {section} (strength {strength})")
#             continue

#         selected_room = suitable_rooms[0]  # pick the smallest suitable room
#         used_rooms.add(selected_room)

#         # Update Firebase
#         institute_ref.child(f"sections/{section}/allocations").update({
#             "classroom": selected_room
#         })

#         print(f"✅ Section {section} (strength {strength}) → {selected_room}")


# Allocating Faculties to each section

from collections import defaultdict
import random
from firebase_admin import db

def allocate_faculty_to_sections(institute):
    # Get references
    faculty_regular_ref = db.reference(f"{institute}/faculty/regular_subjects_faculty")
    faculty_elective_ref = db.reference(f"{institute}/faculty/electives_faculty")
    subjects_regular_ref = db.reference(f"{institute}/subjects/regular_subjects")
    subjects_elective_ref = db.reference(f"{institute}/subjects/elective_subjects")
    sections_ref = db.reference(f"{institute}/sections")

    # Fetch data
    faculty = faculty_regular_ref.get() or {}
    faculty_elective = faculty_elective_ref.get() or {}
    subjects = subjects_regular_ref.get() or {}
    electives = subjects_elective_ref.get() or {}
    sections = sections_ref.get() or {}

    # Validate data
    if not faculty:
        print("❌ No faculty data found")
        return
    if not sections:
        print("❌ No sections found")
        return

    # Track faculty assignments (separate counts for Theory and Lab)
    faculty_assignments = {
        fid: {
            'Theory': 0,
            'Lab': 0,
            'total': 0  # Total non-elective, non-1-credit assignments
        }
        for fid in faculty.keys()
    }

    # Create faculty-subject-type mapping
    faculty_all = {**faculty, **faculty_elective}
    faculty_map = defaultdict(list)
    for fid, fdata in faculty_all.items():
        subject_name = fdata.get('subject', '')
        faculty_type = fdata.get('type', 'Theory')
        if subject_name:
            faculty_map[subject_name].append({
                'id': fid,
                'name': fdata.get('name', ''),
                'type': faculty_type
            })

    # Process each section
    for section_id in sections:
        allocations = {
            'regular_subjects': {},
            'elective_subjects': {}
        }

        # Allocate regular subjects
        for subj_code, subj_info in subjects.items():
            if subj_info.get('elective', False):
                continue  # Skip electives

            subject_name = subj_info['subject']
            subject_type = subj_info.get('type', 'Theory')
            credits = subj_info.get('credits', 0)

            # Skip 1-credit subjects for assignment limits
            apply_limit = credits > 1

            if subject_name in faculty_map:
                eligible_faculty = []
                for fac in faculty_map[subject_name]:
                    # Check type compatibility
                    type_ok = (
                        (fac['type'] == 'Theory & Lab') or
                        (fac['type'] == subject_type)
                    )

                    # Check assignment limits
                    limit_ok = True
                    if apply_limit:
                        current_assignments = faculty_assignments[fac['id']]
                        type_count = current_assignments[subject_type]
                        total_count = current_assignments['total']

                        limit_ok = (type_count < 4) and (total_count < 4)

                    if type_ok and limit_ok:
                        eligible_faculty.append(fac)

                if eligible_faculty:
                    # Select faculty with least assignments
                    selected = min(
                        eligible_faculty,
                        key=lambda x: faculty_assignments[x['id']]['total']
                    )

                    allocations['regular_subjects'][subj_code] = {
                        'faculty_id': selected['id'],
                        'faculty_name': selected['name'],
                        'subject_name': subj_info['subject'],
                        'type': subject_type,
                        'credits': credits
                    }

                    # Update assignment counts
                    if apply_limit:
                        faculty_assignments[selected['id']][subject_type] += 1
                        faculty_assignments[selected['id']]['total'] += 1

        # Allocate elective subjects (no limits)
        for elec_code, elec_info in electives.items():
            subject_name = elec_info['subject']
            subject_type = elec_info.get('type', 'Theory')

            if subject_name in faculty_map:
                eligible_faculty = [
                    fac for fac in faculty_map[subject_name]
                    if (fac['type'] == 'Theory & Lab') or (fac['type'] == subject_type)
                ]

                if eligible_faculty:
                    selected = random.choice(eligible_faculty)
                    allocations['elective_subjects'][elec_code] = {
                        'faculty_id': selected['id'],
                        'faculty_name': selected['name'],
                        'subject_name': elec_info['subject'],
                        'type': subject_type
                    }

        # Update database
        sections_ref.child(f"{section_id}/allocations").update(allocations)
        print(f"✅ Updated allocations for section {section_id}")

    print("🎉 Faculty allocation completed with all constraints!")


# Creating time slots

from datetime import datetime, timedelta

def time_slots(start_time, end_time, duration, breaks, breaks_duration):
    slots = []
    st = start_time
    et = st + timedelta(minutes=duration)
    counter = 0
    breaks_inserted = 0

    # We'll keep a small buffer of generated slots to insert breaks "before the last class"
    buffer_slots = []

    while et <= end_time:
        # Add the current class slot to buffer
        buffer_slots.append((st, et))
        counter += 1

        # After 2 classes, insert first break immediately
        if counter == 2 and breaks_inserted < breaks:
            break_st = et
            break_et = break_st + timedelta(minutes=breaks_duration)
            buffer_slots.append((break_st, break_et))
            breaks_inserted += 1
            st = break_et
        else:
            st = et

        et = st + timedelta(minutes=duration)

    # Now we have all classes and the first break in buffer_slots
    # Insert the last break before the last class if breaks >= 2 and not already inserted
    if breaks >= 2 and breaks_inserted < breaks and len(buffer_slots) >= 2:
        # The last class is the last slot if no break at end, or second last if break present
        # Insert break before last class slot
        last_class = buffer_slots.pop()  # remove last class temporarily

        # Insert break before last class
        last_class_start = last_class[0]
        break_st = last_class_start - timedelta(minutes=breaks_duration)
        break_et = last_class_start

        # Check if break_st >= previous slot end to avoid overlap
        if buffer_slots[-1][1] <= break_st:
            buffer_slots.append((break_st, break_et))
        else:
            # If overlap, adjust break start to last slot end (silent shift)
            break_st = buffer_slots[-1][1]
            break_et = break_st + timedelta(minutes=breaks_duration)
            # shift last class start by break duration
            last_class = (break_et, last_class[1] + timedelta(minutes=breaks_duration))
            buffer_slots.append((break_st, break_et))

        breaks_inserted += 1
        buffer_slots.append(last_class)

    # If more breaks remain, insert them silently between classes, avoiding consecutive breaks
    remaining_breaks = breaks - breaks_inserted
    if remaining_breaks > 0:
        # Insert remaining breaks between classes with spacing
        i = 0
        while remaining_breaks > 0 and i < len(buffer_slots) - 1:
            # Only insert breaks between two class slots, avoid if next slot is break already
            curr_slot = buffer_slots[i]
            next_slot = buffer_slots[i+1]

            # Ensure current and next slots are classes (not breaks)
            # Let's assume breaks are slots of length breaks_duration, classes are duration

            curr_length = (curr_slot[1] - curr_slot[0]).total_seconds() / 60
            next_length = (next_slot[1] - next_slot[0]).total_seconds() / 60

            if curr_length == duration and next_length == duration:
                # Insert break after current slot if no break already after it
                break_st = curr_slot[1]
                break_et = break_st + timedelta(minutes=breaks_duration)

                # Insert break if it doesn't overlap with next slot start
                if break_et <= next_slot[0]:
                    buffer_slots.insert(i + 1, (break_st, break_et))
                    remaining_breaks -= 1
                    i += 2  # Skip over inserted break and next slot
                else:
                    i += 1
            else:
                i += 1

    return buffer_slots


# Updating Firebase with time slots and timetable in sections and venues

from datetime import datetime
from firebase_admin import db

def upload_class_timetable_to_firebase(institute, start_str, end_str, duration, breaks, break_duration, num_days):
    # Step 1: Convert time strings to datetime objects
    start_time = datetime.strptime(start_str, "%H:%M")
    end_time = datetime.strptime(end_str, "%H:%M")

    # Step 2: Define weekdays based on number of working days
    weekdays = ["1_Monday", "2_Tuesday", "3_Wednesday", "4_Thursday", "5_Friday"]
    if num_days == 6:
        weekdays.append("6_Saturday")

    # Step 3: Fetch all sections
    institute_ref = db.reference(f"{institute}")
    sections_data = institute_ref.child("sections").get()

    if not sections_data:
        print("No sections found under the institute.")
        return

    for section in sections_data:
        timetable_ref = institute_ref.child(f"sections/{section}/timetable")

        for day in weekdays:
            # Step 4: Call your existing time_slots() function
            slots = time_slots(start_time, end_time, duration, breaks, break_duration)

            # Step 5: Filter out breaks (where duration == break_duration)
            class_slots = [
                (slot[0], slot[1])
                for slot in slots
                if int((slot[1] - slot[0]).total_seconds() / 60) != break_duration
            ]

            # Step 6: Upload each class slot to Firebase with default type = "nill"
            day_ref = timetable_ref.child(day)
            for i, (st, et) in enumerate(class_slots, 1):
                period_data = {
                    "start": st.strftime("%H:%M"),
                    "end": et.strftime("%H:%M"),
                    "type": "nill"
                }
                day_ref.child(f"Period {i}").set(period_data)

    print(f"✅ Timetable initiated successfully for all sections")

  #  upload_venue_timetable_to_firebase(institute, start_str, end_str, duration, breaks, break_duration, num_days)


from datetime import datetime
from firebase_admin import db

def upload_venue_timetable_to_firebase(institute, start_str, end_str, duration, breaks, break_duration, num_days):
    # Step 1: Convert time strings to datetime objects
    start_time = datetime.strptime(start_str, "%H:%M")
    end_time = datetime.strptime(end_str, "%H:%M")

    # Step 2: Define weekdays
    weekdays = ["1_Monday", "2_Tuesday", "3_Wednesday", "4_Thursday", "5_Friday"]
    if num_days == 6:
        weekdays.append("6_Saturday")

    # Step 3: Fetch all venues
    institute_ref = db.reference(f"{institute}")
    venues_data = institute_ref.child("venues").get()

    if not venues_data:
        print("No venues found under the institute.")
        return

    for venue in venues_data:
        timetable_ref = institute_ref.child(f"venues/{venue}/timetable")

        for day in weekdays:
            # Step 4: Generate time slots
            slots = time_slots(start_time, end_time, duration, breaks, break_duration)

            # Step 5: Filter out breaks
            class_slots = [
                (slot[0], slot[1])
                for slot in slots
                if int((slot[1] - slot[0]).total_seconds() / 60) != break_duration
            ]

            # Step 6: Initialize venue timetable for each slot
            day_ref = timetable_ref.child(day)
            for i, (st, et) in enumerate(class_slots, 1):
                period_data = {
                    "start": st.strftime("%H:%M"),
                    "end": et.strftime("%H:%M"),
                    "bookings": []  # Multiple bookings allowed
                }
                day_ref.child(f"Period {i}").set(period_data)

    print(f"✅ Venue timetable initialized with support for multiple section bookings.")
 


# Assigning Lunch breaks

from datetime import datetime
import firebase_admin
from firebase_admin import credentials, db

def is_within_time_window(start_str, target_hour):
    try:
        start_time = datetime.strptime(start_str, "%H:%M")
        target_time = start_time.replace(hour=target_hour, minute=0)
        delta = abs((start_time - target_time).total_seconds()) / 60
        return delta <= 15
    except:
        return False

def assign_lunch_breaks_in_firebase(institute):
    import random

    sections_ref = db.reference(f"{institute}/sections")
    sections_data = sections_ref.get()

    if not sections_data:
        print("❌ No section data found.")
        return

    section_names = sorted(sections_data.keys())
    missing_after_pass = []

    for section in section_names:
        timetable_ref = db.reference(f"{institute}/sections/{section}/timetable")
        timetable = timetable_ref.get()

        if not timetable:
            print(f"⚠️ No timetable found for {section}")
            continue

        for day, slots in timetable.items():
            # Skip if lunch already exists
            if any(slot.get('type') == "Lunch" for slot in slots.values()):
                continue

            # First: try 12–1 PM window
            eligible_slots = [
                k for k, v in slots.items()
                if v.get("type") == "nill" and (
                    is_within_time_window(v.get("start", ""), 12) or is_within_time_window(v.get("start", ""), 13)
                )
            ]

            if eligible_slots:
                chosen = random.choice(eligible_slots)
                slots[chosen]["type"] = "Lunch"
            else:
                # Second: try any available nill slot
                fallback_slots = [
                    k for k, v in slots.items() if v.get("type") == "nill"
                ]
                if fallback_slots:
                    chosen = random.choice(fallback_slots)
                    slots[chosen]["type"] = "Lunch"
                else:
                    missing_after_pass.append((section, day))

            timetable_ref.child(day).set(slots)

    print("🍱 Lunch breaks assigned with double-check and fallback.")
    if missing_after_pass:
        print("⚠️ These section-days had no available slot to assign lunch:")
        for s, d in missing_after_pass:
            print(f" - Section {s}, Day {d}")
    else:
        print("✅ Every day in every section has a lunch period.")



# Alloting Electives to time slots

import firebase_admin
from firebase_admin import db
import random
from collections import defaultdict

def assign_electives_to_sections(institute):
    """
    Assigns 2 continuous elective periods on 2 different days for each section.
    Ensures:
    - No slot marked as 'Lunch' or non-'nill' is overwritten.
    - No section has more than 2 elective periods in a single day.
    """
    import random
    from firebase_admin import db

    sections_ref = db.reference(f"{institute}/sections")
    sections = sections_ref.get()

    if not sections:
        print("❌ No sections found.")
        return

    section_ids = sorted(sections.keys())
    total = len(section_ids)

    any_section = next(iter(sections.values()))
    days = sorted(any_section['timetable'].keys())
    periods = sorted(any_section['timetable'][days[0]].keys())

    # Generate valid continuous period pairs (e.g., P1 & P2)
    period_pairs = [(periods[i], periods[i + 1]) for i in range(len(periods) - 1)]
    slot_combos = [(d, pp) for d in days for pp in period_pairs]
    random.shuffle(slot_combos)

    num_groups = 3
    sections_per_group = total // num_groups

    for group_index in range(num_groups):
        group_start = group_index * sections_per_group
        group_end = (group_index + 1) * sections_per_group if group_index < num_groups - 1 else total
        group_sections = section_ids[group_start:group_end]

        valid_combos = []
        daily_counts = {sec_id: {} for sec_id in group_sections}

        for d, (p1, p2) in slot_combos:
            valid = True
            for sec_id in group_sections:
                slots = sections[sec_id]['timetable'][d]
                # Check if already too many electives on this day
                if daily_counts[sec_id].get(d, 0) >= 2:
                    valid = False
                    break
                # Check if both periods are available
                if slots[p1].get('type') != 'nill' or slots[p2].get('type') != 'nill':
                    valid = False
                    break
            if valid:
                valid_combos.append((d, (p1, p2)))
                for sec_id in group_sections:
                    daily_counts[sec_id][d] = daily_counts[sec_id].get(d, 0) + 2
            if len(valid_combos) >= 2:
                break

        if len(valid_combos) < 2:
            print(f"⚠️ Couldn't find 2 separate elective slots for group {group_index + 1}")
            continue

        for sec_id in group_sections:
            for d, (p1, p2) in valid_combos:
                sec_ref = sections_ref.child(f"{sec_id}/timetable/{d}")
                sec_ref.child(p1).update({"type": "Elective"})
                sec_ref.child(p2).update({"type": "Elective"})

    print("🎓 Electives assigned safely with max 2 elective periods per day per section.")



# Assigning Theory Classes

def assign_theory_periods(institute):
    import random
    from firebase_admin import db

    section_ref = db.reference(f"{institute}/sections")
    subject_ref = db.reference(f"{institute}/subjects/regular_subjects")
    venue_ref = db.reference(f"{institute}/venues")

    timetable_data = section_ref.get()
    subjects_data = subject_ref.get()
    venues_data = venue_ref.get()

    # Filter valid classrooms
    classroom_venues = {
        venue_name: details.get("capacity", 0)
        for venue_name, details in venues_data.items()
        if details.get("type") == "Classroom"
    }

    if not classroom_venues:
        raise ValueError("No 'Classroom' type venues found in database.")

    for section_key, section in timetable_data.items():
        timetable = section.get("timetable", {})
        allocations = section.get("allocations", {}).get("regular_subjects", {})
        if not timetable:
            continue

        section_strength = section.get("strength", 0)
        preferred_venue = section.get("venue")

        def venue_fits(v):
            return classroom_venues.get(v, 0) >= section_strength

        assigned_venue_valid = preferred_venue and venue_fits(preferred_venue)

        for subj_key, subject in subjects_data.items():
            if subject.get("type") != "Theory":
                continue

            credits = subject.get("credits", 0)
            faculty_info = allocations.get(subj_key, {})

            while credits > 0:
                day = random.choice(list(timetable.keys()))
                periods = timetable[day]
                if not isinstance(periods, dict):
                    continue

                assigned_count = sum(
                    1 for p in periods.values() if p.get("subject") == subj_key
                )
                if assigned_count >= 2:
                    continue

                period_keys = sorted(periods.keys(), key=lambda x: int(x.split()[-1]))
                assigned = False

                for i, p_key in enumerate(period_keys):
                    current_period = periods[p_key]
                    if current_period.get("type") != "nill":
                        continue

                    def pick_venue():
                        if assigned_venue_valid and random.random() < 0.8:
                            return preferred_venue
                        else:
                            alternatives = [
                                v for v in classroom_venues if venue_fits(v)
                            ]
                            if preferred_venue in alternatives:
                                alternatives.remove(preferred_venue)
                            return random.choice(alternatives) if alternatives else None

                    # Double Period Assignment
                    if credits >= 2 and i + 1 < len(period_keys):
                        next_period = periods[period_keys[i + 1]]
                        if next_period.get("type") == "nill" and assigned_count == 0:
                            venue = pick_venue()
                            if not venue:
                                continue

                            for j, period in zip([i, i + 1], (current_period, next_period)):
                                period_name = period_keys[j]
                                start = period.get("start")
                                end = period.get("end")

                                # Update section timetable
                                period.update({
                                    "type": "regular",
                                    "subject": subj_key,
                                    "venue": venue
                                })

                                # Mirror in venue timetable
                                venue_period_ref = db.reference(f"{institute}/venues/{venue}/timetable/{day}/{period_name}")
                                venue_period_ref.set({
                                "start": start,
                                "end": end,
                                "bookings": {
                                    section_key: {
                                        "subject": subj_key,
                                        "section": section_key,
                                        "faculty_subject": faculty_info.get("faculty_subject"),
                                        "faculty": faculty_info.get("faculty_name")
                                    }
                                }
                            })


                            credits -= 2
                            assigned = True
                            break

                    # Single Period Assignment
                    if credits > 0:
                        venue = pick_venue()
                        if not venue:
                            continue

                        start = current_period.get("start")
                        end = current_period.get("end")

                        # Update section timetable
                        current_period.update({
                            "type": "regular",
                            "subject": subj_key,
                            "venue": venue
                        })

                        # Mirror in venue timetable
                        venue_period_ref = db.reference(f"{institute}/venues/{venue}/timetable/{day}/{p_key}")
                        venue_period_ref.set({
                        "start": start,
                        "end": end,
                        "bookings": {
                            section_key: {
                                "subject": subj_key,
                                "section": section_key,
                                "faculty_subject": faculty_info.get("faculty_subject"),
                                "faculty": faculty_info.get("faculty_name")
                            }
    }
                    })



                        credits -= 1
                        assigned = True
                        break

                if not assigned:
                    break

        # Push updated section timetable
        section_ref.child(f"{section_key}/timetable").set(timetable)



from collections import defaultdict

def detect_clashes(institute, max_iterations=3):
    section_ref = db.reference(f"{institute}/sections")
    venue_ref = db.reference(f"{institute}/venues")
    
    sections = section_ref.get()
    venues = venue_ref.get()

    if not sections or not venues:
        print("❌ Missing sections or venues.")
        return

    classroom_venues = {
        name: v["capacity"]
        for name, v in venues.items()
        if v.get("type") == "Classroom"
    }

    iteration = 0
    while iteration < max_iterations:
        iteration += 1
        print(f"\n🔍 Iteration {iteration}: Checking for timetable clashes...")

        schedule_map = defaultdict(lambda: {"venues": {}, "faculty": {}})
        time_venue_usage = defaultdict(lambda: defaultdict(list))  # time → venue → [section]
        faculty_subject_grouping = defaultdict(list)

        clash_found = False

        for sec_id, sec_data in sections.items():
            timetable = sec_data.get("timetable", {})
            sec_strength = sec_data.get("strength", 0)

            for day, periods in timetable.items():
                for period_id, p in periods.items():
                    if p.get("type") in ["nill", "Lunch"]:
                        continue

                    start = p.get("start")
                    end = p.get("end")
                    venue = p.get("venue")
                    faculty = p.get("faculty_id") or p.get("faculty")
                    subject = p.get("subject")

                    key = f"{day}|{start}|{end}"
                    time_venue_usage[key][venue].append((sec_id, sec_strength))
                    
                    # Group for merge logic
                    if subject and faculty:
                        merge_key = f"{day}|{start}|{end}|{faculty}|{subject}"
                        faculty_subject_grouping[merge_key].append({
                            "section": sec_id,
                            "strength": sec_strength,
                            "period_id": period_id,
                            "venue": venue
                        })

        # Handle faculty-subject clashes with combined strength
        for key, group in faculty_subject_grouping.items():
            if len(group) <= 1:
                continue  # No merging necessary

            day, start, end, faculty, subject = key.split("|")
            total_strength = sum(entry["strength"] for entry in group)
            common_venue = group[0]["venue"]
            venue_capacity = classroom_venues.get(common_venue, 0)

            if total_strength > venue_capacity:
                clash_found = True
                print(f"🚨 Capacity Clash: {faculty} teaching {subject} to sections {[e['section'] for e in group]} in '{common_venue}' exceeds capacity ({total_strength} > {venue_capacity})")

                # Try to find alternative venue
                new_venue = find_available_venue(day, start, end, classroom_venues, time_venue_usage, total_strength)

                if new_venue:
                    print(f"🔁 Reassigning merged group to '{new_venue}'")
                    for entry in group:
                        section_ref.child(f"{entry['section']}/timetable/{day}/{entry['period_id']}/venue").set(new_venue)
                        time_venue_usage[f"{day}|{start}|{end}"][new_venue].append((entry["section"], entry["strength"]))
                else:
                    print(f"⚠️ No suitable venue found for group: {group}")

        if clash_found:
            print("⚠️ Clashes detected. Attempting to resolve...\n")
            sections = section_ref.get()  # Refresh data after resolution
        else:
            print("✅ No clashes found.")
            return

def find_available_venue(day, start, end, classroom_venues, time_venue_usage, required_capacity):
    time_key = f"{day}|{start}|{end}"
    used_venues = set(time_venue_usage[time_key].keys())

    for vname, cap in classroom_venues.items():
        if cap >= required_capacity and vname not in used_venues:
            return vname
    return None


def resolve_venue_clashes(institute):
    print("🔧 Resolving venue clashes...")

    section_ref = db.reference(f"{institute}/sections")
    venue_ref = db.reference(f"{institute}/venues")

    sections = section_ref.get()
    venues = venue_ref.get()

    if not sections or not venues:
        print("❌ Missing sections or venues.")
        return

    # Get all classroom-type venues with their capacities
    classroom_venues = {
        name: v["capacity"]
        for name, v in venues.items()
        if v.get("type") == "Classroom"
    }

    # Build a time → venue → [entry list] map
    time_map = defaultdict(lambda: defaultdict(list))  # time_key → venue → [entries]
    faculty_subject_map = defaultdict(list)  # (time_key, faculty, subject) → [entries]

    for sec_id, sec_data in sections.items():
        timetable = sec_data.get("timetable", {})
        sec_strength = sec_data.get("strength", 0)

        for day, periods in timetable.items():
            for period_id, p in periods.items():
                if p.get("type") in ["nill", "Lunch"]:
                    continue

                start, end = p.get("start"), p.get("end")
                venue = p.get("venue")
                faculty = p.get("faculty_id") or p.get("faculty")
                subject = p.get("subject")

                if not all([venue, start, end, faculty, subject]):
                    continue

                time_key = f"{day}|{start}|{end}"
                entry = {
                    "section": sec_id,
                    "faculty": faculty,
                    "subject": subject,
                    "period_id": period_id,
                    "venue": venue,
                    "strength": sec_strength
                }

                time_map[time_key][venue].append(entry)
                faculty_subject_map[(time_key, faculty, subject)].append(entry)

    # ✅ Resolve regular venue clashes
    for time_key, venue_data in time_map.items():
        for venue, entries in venue_data.items():
            if len(entries) <= 1:
                continue  # No conflict

            print(f"\n⏰ Time Slot: {time_key}")
            print(f"🏫 Venue Clash in '{venue}' between sections: {[e['section'] for e in entries]}")

            used_venues = {venue}
            for i, entry in enumerate(entries):
                if i == 0:
                    print(f"✔️ Keeping Section {entry['section']} in '{venue}'")
                    continue

                new_venue = find_alternate_venue(classroom_venues, used_venues, time_map[time_key], entry["strength"])
                if new_venue:
                    day, start, end = time_key.split('|')
                    print(f"🔁 Reassigning Section {entry['section']} → '{new_venue}' at {day} {start}-{end}")
                    section_ref.child(f"{entry['section']}/timetable/{day}/{entry['period_id']}/venue").set(new_venue)
                    used_venues.add(new_venue)
                    time_map[time_key][new_venue].append(entry)
                else:
                    print(f"⚠️ Could not find alternative venue for Section {entry['section']} (strength {entry['strength']}) at {time_key}")

    # ✅ Resolve same faculty+subject + combined strength > venue capacity
    for (time_key, faculty, subject), entries in faculty_subject_map.items():
        if len(entries) <= 1:
            continue

        same_venue = all(e["venue"] == entries[0]["venue"] for e in entries)
        combined_strength = sum(e["strength"] for e in entries)
        venue = entries[0]["venue"]
        cap = classroom_venues.get(venue, 0)

        if same_venue and combined_strength > cap:
            print(f"\n⚠️ Capacity Clash for '{faculty}' teaching '{subject}' at {time_key}")
            print(f"📚 Sections: {[e['section'] for e in entries]} in '{venue}' (Combined: {combined_strength} > Capacity: {cap})")

            used_venues = set(time_map[time_key].keys())  # venues already used at this time
            new_venue = find_alternate_venue(classroom_venues, used_venues, time_map[time_key], combined_strength)

            if new_venue:
                day, start, end = time_key.split('|')
                print(f"🔁 Reassigning ALL to '{new_venue}' at {day} {start}-{end}")
                for entry in entries:
                    section_ref.child(f"{entry['section']}/timetable/{day}/{entry['period_id']}/venue").set(new_venue)
                    time_map[time_key][new_venue].append(entry)
            else:
                print(f"❌ No large enough venue available for combined session ({combined_strength}) at {time_key}")

    print("\n✅ Venue clash resolution complete.\n")


def find_alternate_venue(classroom_venues, used_venues, current_venue_map, required_capacity):
    for vname, cap in classroom_venues.items():
        if vname not in used_venues and vname not in current_venue_map and cap >= required_capacity:
            return vname
    return None



# Generating Excel Files
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import db

def generate_excel_reports(institute, output_file="./All_Sections_Report.xlsx"):
    """
    Generates a single Excel file with sheets for each section.
    Each sheet contains the timetable followed by the faculty allocation.
    """
    try:
        db_ref = db.reference()

        sections = db_ref.child(f"{institute}/sections").get() or {}
        faculty_data = db_ref.child(f"{institute}/faculty").get() or {}

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sec_id, sec_data in sections.items():
                if not isinstance(sec_data, dict):
                    print(f"⚠️ Invalid section data for {sec_id}")
                    continue

                timetable = sec_data.get("timetable", {})
                allocations = sec_data.get("allocations", {})
                regular_subjects = allocations.get("regular_subjects", {})
                elective_subjects = allocations.get("elective_subjects", {})

                # ========== TIMETABLE SHEET ==========
                period_data = {}
                weekdays = []

                for day_key, day_schedule in timetable.items():
                    if not isinstance(day_schedule, dict):
                        continue

                    try:
                        day_parts = day_key.split('_')
                        day_num = int(day_parts[0])
                        day_name = '_'.join(day_parts[1:]) if len(day_parts) > 1 else day_key
                        weekdays.append((day_num, day_name, day_key))
                    except (ValueError, IndexError):
                        continue

                    if not period_data:
                        for period, slot in day_schedule.items():
                            if isinstance(slot, dict):
                                period_data[period] = (
                                    slot.get("start", ""),
                                    slot.get("end", "")
                                )

                weekdays.sort()
                sorted_days = [day[2] for day in weekdays]

                timetable_rows = []
                for day_key in sorted_days:
                    day_schedule = timetable.get(day_key, {})
                    day_name = day_key.split('_')[-1] if '_' in day_key else day_key
                    row = [day_name]

                    for period in sorted(period_data.keys()):
                        slot = day_schedule.get(period, {})
                        if not isinstance(slot, dict):
                            row.append("")
                            continue

                        period_type = slot.get("type", "nill")

                        if period_type == "nill":
                            row.append("")
                        elif period_type == "Lunch":
                            row.append("Lunch")
                        else:
                            subject = slot.get("subject", "")
                            venue = slot.get("venue", "")
                            if subject:
                                row.append(f"{subject} ({venue})" if venue else subject)
                            else:
                                row.append(period_type.capitalize())

                    timetable_rows.append(row)

                column_headers = [
                    f"{start}-{end}" if start and end else period 
                    for period, (start, end) in sorted(period_data.items())
                ]
                timetable_df = pd.DataFrame(
                    timetable_rows, 
                    columns=["Day"] + column_headers
                ) if timetable_rows else pd.DataFrame(["No timetable data available"])

                # ========== FACULTY SHEET ==========
                faculty_rows = []

                if isinstance(regular_subjects, dict):
                    for subj_id, info in regular_subjects.items():
                        if not isinstance(info, dict):
                            continue
                        faculty_id = info.get("faculty_id")
                        faculty_name = (
                            faculty_data.get('regular_subjects_faculty', {})
                            .get(faculty_id, {})
                            .get("name", "Unassigned")
                        )
                        faculty_rows.append([faculty_name, subj_id])

                if isinstance(elective_subjects, dict):
                    for subj_id, info in elective_subjects.items():
                        if not isinstance(info, dict):
                            continue
                        faculty_id = info.get("faculty_id")
                        faculty_name = (
                            faculty_data.get('electives_faculty', {})
                            .get(faculty_id, {})
                            .get("name", "Unassigned")
                        )
                        faculty_rows.append([faculty_name, subj_id])

                faculty_df = pd.DataFrame(
                    faculty_rows, 
                    columns=["Faculty", "Subject"]
                ) if faculty_rows else pd.DataFrame(["No faculty assignments available"])

                # ========== WRITE TO SHEET ==========
                sheet_name = sec_id[:31]  # Excel limits sheet name to 31 characters

                timetable_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                faculty_start_row = len(timetable_df) + 4  # 4 empty rows after timetable
                faculty_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=faculty_start_row)

        # ========== AUTO-ADJUST COLUMN WIDTHS ==========
        wb = load_workbook(output_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        wb.save(output_file)

        print(f"✅ Master Excel report generated: {output_file}")

    except Exception as e:
        print(f"❌ Error generating Excel report: {str(e)}")
        raise
        

# main function

import random
from firebase_admin import db

def main():

    institute = input("Enter the institute Firebase node name: ").strip()
    start_time_str = input("Enter start time of classes (HH:MM, e.g., 08:00): ").strip()
    end_time_str = input("Enter end time of classes (HH:MM, e.g., 16:00): ").strip()

    try:
        period_duration = int(input("Enter duration of each class period (in minutes, e.g., 55): ").strip())
        num_breaks = int(input("Enter number of short breaks in a day (e.g., 2): ").strip())
        break_duration = int(input("Enter duration of each short break (in minutes, e.g., 20): ").strip())
        num_days = int(input("Enter number of working days (5 for Mon-Fri, 6 for Mon-Sat): ").strip())

        if num_days not in [5, 6]:
            print("Invalid number of days. Please enter 5 or 6.")
            return

    except ValueError:
        print("Invalid numeric input. Please enter valid integers.")
        return


    # Call the upload function
    upload_university_data(institute)
    upload_class_timetable_to_firebase(
        institute=institute,
        start_str=start_time_str,
        end_str=end_time_str,
        duration=period_duration,
        breaks=num_breaks,
        break_duration=break_duration,
        num_days=num_days
    )
    upload_venue_timetable_to_firebase(institute, start_time_str, end_time_str, period_duration, num_breaks, break_duration, num_days)
    metadata_ref = db.reference(f"{institute}/metadata")
    if not metadata_ref.child("lunch_assigned").get():
        assign_lunch_breaks_in_firebase(institute)
        metadata_ref.child("lunch_assigned").set(True)
        print("Done")
    else:
        print("✔️ Lunch breaks already assigned. Skipping.")

    # Scheduling Classes
    assign_electives_to_sections(institute)
    assign_theory_periods(institute)
    detect_clashes(institute, max_iterations=3)
    generate_excel_reports(institute, output_file="./All_Sections_Report.xlsx")

# Run the main function
if __name__ == "__main__":
    main()